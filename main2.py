import json
import random
import time
from selenium import webdriver
from selenium_stealth import stealth
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import requests
from bs4 import BeautifulSoup
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side
from openpyxl import load_workbook
from datetime import datetime
from aiogram import types, executor, Dispatcher, Bot
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.dispatcher import FSMContext


headers = {
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.127 Safari/537.36",
    "accept": "application/json, text/plain, */*"
}


def selenium_nou_bot(url):
    # "Забирает Html код"
    options = webdriver.ChromeOptions()
    options.add_argument("--no-sandbox")
    options.add_argument("start-maximized")
    # options.add_argument('--proxy-server=46.3.182.109:8000')
    options.add_argument("--headless")

    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    # старый метод
    # driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
    # новый метод
    s = Service(executable_path=ChromeDriverManager().install())
    driver = webdriver.Chrome(service=s, options=options)

    stealth(driver,
            languages=["en-US", "en"],
            vendor="Google Inc.",
            platform="Win32",
            webgl_vendor="Intel Inc.",
            renderer="Intel Iris OpenGL Engine",
            fix_hairline=True,
            )

    url = url
    time.sleep(random.randint(5, 8))
    try:
        driver.get(url)
    except:
        print("Неверная ссыка")
        return


    if not os.path.exists("data"):
        os.mkdir("data")

    offset = 0
    # прокручиваем 5 раз, это 100 карточек
    while True:
        if offset == 11:
            break
        else:
            offset += 1
            time.sleep(4)
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")


    # сохраним весь главный сайт
    # сохраним весь главный сайт
    with open(f"data/index.html", "w", encoding='utf-8') as file:
        file.write(driver.page_source)

    # закроем браузер
    driver.quit()

    # откроем сохраненный сайт
    with open(f"data/index.html", encoding="utf-8") as file:
        src = file.read()

    soup = BeautifulSoup(src, 'lxml')

    return soup


def html_cads(url, project_name):
    global driver
    # "Забирает Html код"
    options = webdriver.ChromeOptions()
    options.add_argument("--no-sandbox")
    options.add_argument("start-maximized")
    # options.add_argument('--proxy-server=46.3.182.109:8000')
    options.add_argument("--headless")

    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    # старый метод
    # driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
    # новый метод
    s = Service(executable_path=ChromeDriverManager().install())
    driver = webdriver.Chrome(service=s, options=options)

    stealth(driver,
            languages=["en-US", "en"],
            vendor="Google Inc.",
            platform="Win32",
            webgl_vendor="Intel Inc.",
            renderer="Intel Iris OpenGL Engine",
            fix_hairline=True,
            )

    url = url
    time.sleep(random.randint(5, 7))
    driver.get(url)



    # сохраним файл под именем url на карту
    with open(f"data/{project_name}.html", "w", encoding='utf-8') as file:
        file.write(driver.page_source)


    # откроем сохраненный файл
    with open(f"data/{project_name}.html", encoding="utf-8") as file:
        src = file.read()

    return src




def excel_loader(row, title, price, result_offers):
    # global row
    """Сохраняем данные в excel"""
    try:
        # если файл есть дописываем
        book = openpyxl.load_workbook("my_book.xlsx")
    except:
        # если нет создаем
        book = openpyxl.Workbook()

    # active берет первый лист
    sheet = book.active

    # создадим заголовки
    sheet['A1'] = title_magic
    sheet['B1'] = 'Цена'
    sheet['C1'] = 'Предложения'

    # начинаем поиск пустого ряда с ряда номер 2

    # if sheet[f'A{row}'].value is None:
    sheet[f'A{row}'].value = title
    sheet[f'B{row}'].value = price
    sheet[f'C{row}'].value = result_offers
    row += 1
    print("Работаю")
    book.save("my_book.xlsx")
    book.close()




TOKEN = "your token from bot telegram"

storage = MemoryStorage()

bot = Bot(token=TOKEN)
dp = Dispatcher(bot, storage=storage)


async def on_startup(_):
    print("Бот вышел в онлайн")
    # передать в executor


@dp.message_handler(commands='start')
async def begin(message: types.Message):
    await bot.send_message(message.chat.id, "Пришли ссылку")




class FSMAdmin(StatesGroup):
    otv = State()


@dp.message_handler(content_types=['text'], state=None)
async def text(message: types.Message):
    global sylka
    sylka = message.text
    await bot.send_message(message.chat.id, "Введите наименование")
    await FSMAdmin.next()



@dp.message_handler(state=FSMAdmin.otv)
async def otver(message: types.Message, state: FSMContext):
    global title_magic
    title_magic = message.text
    await bot.send_message(message.chat.id, "Парсинг запущен! я сообщу вам когда файл будет готов")
    await FSMAdmin.next()
    #____________________________________________________
    def get_data_file(headers):
        row = 2
        # global driver, row
        soup = selenium_nou_bot(sylka)

        try:
            urls = soup.find_all("div", class_='tw-rounded-xl tw-overflow-hidden me-flex-center')
        except:
            return "Неверная ссылка!!!"
        # print(urls)

        projects_urls = []
        for url in urls:
            project_url = "https://magiceden.io" + url.find('a').get('href')
            projects_urls.append(project_url)

        # для теста спарсим только два карточки
        # rows = len(projects_urls)
        number_name = 1
        for project_url in projects_urls:
            print(f"Захожу на страницу: {number_name}", project_url)
            # забираем название из url
            project_name = project_url.split("/")[4].strip()
            src = html_cads(url=project_url, project_name=number_name)
            number_name += 1

            # забираем данные с карточек
            soup = BeautifulSoup(src, 'lxml')
            time.sleep(4)

            title = soup.find("h3", class_="tw-m-0 item-title tw-text-xl")
            if title:
                title = title.text

                price = soup.find("span", class_="text-white price")

                if price:
                    price = price.text

                try:
                    offers_price = soup.find("tbody", role="rowgroup").find("span", class_="tw-text-sm").text
                    # offers = soup.find("tbody", role="rowgroup").find("tr", role="row").text.split(" ")[1]
                    # result_offers = offers_price + " " + offers
                    result_offers = offers_price
                except Exception as ex:
                    result_offers = "Нет"

                finally:
                    excel_loader(row, title, price, result_offers)
                    row += 1

            # закроем браузер
            # driver.quit()

            # print(title)
            # print(result_offers)
        return "Сбор данных успешно завершен, можете забрать файл с данными с сервера!"

    otvet = get_data_file(headers=headers)


    #____________________________________________________

    await bot.send_message(message.chat.id, otvet)

    await state.finish()




executor.start_polling(dp, skip_updates=True, on_startup=on_startup)